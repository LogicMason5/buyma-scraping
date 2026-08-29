"""Buyma listing row builder + delivery Excel (sample 納品シート format)."""

from __future__ import annotations

import csv
import logging
import math
import re
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from slugify import slugify

from core.config import get_settings
from core.prompts.defaults import BUYMA_COLOR_SIZE_NOTE_TEMPLATE, BUYMA_PRODUCT_COMMENT_TEMPLATE

logger = logging.getLogger(__name__)

_delivery_lock = threading.Lock()

BUYMA_CSV_HEADERS = [
    "フォルダ名",
    "商品名",
    "ブランド",
    "モデル・ライン",
    "カテゴリ",
    "商品コメント",
    "色・サイズ補足情報",
    "購入期限",
    "仕入先URL",
    "買付地",
    "ショップ名",
    "発送地",
    "仕入先保存",
    "カラー系統",
    "サイズ",
    "シーズン",
    "タグ",
    "テーマ",
    "価格",
    "参考価格",
    "配送方法名",
    "在庫",
    "型番/メモ",
    "関税負担",
    "出品メモ",
    "ライバル価格",
    "ライバルURL",
    "SKU",
    "価格計算式",
    "品代(海外仕入)",
    "送料(仕入)",
    "梱包",
    "送料(出品)",
    "出品価格",
    "関税(10%)",
    "消費税(10%)",
    "仕入合計",
    "BUYMA手数料",
    "利益%",
    "出品価格",  # duplicate header as in sample (col AN)
    "利益",
]

# Unique keys for the duplicate 「出品価格」 columns (S/AH vs AN).
_ROW_KEYS = [
    "フォルダ名",
    "商品名",
    "ブランド",
    "モデル・ライン",
    "カテゴリ",
    "商品コメント",
    "色・サイズ補足情報",
    "購入期限",
    "仕入先URL",
    "買付地",
    "ショップ名",
    "発送地",
    "仕入先保存",
    "カラー系統",
    "サイズ",
    "シーズン",
    "タグ",
    "テーマ",
    "価格",
    "参考価格",
    "配送方法名",
    "在庫",
    "型番/メモ",
    "関税負担",
    "出品メモ",
    "ライバル価格",
    "ライバルURL",
    "SKU",
    "価格計算式",
    "品代(海外仕入)",
    "送料(仕入)",
    "梱包",
    "送料(出品)",
    "出品価格",
    "関税(10%)",
    "消費税(10%)",
    "仕入合計",
    "BUYMA手数料",
    "利益%",
    "出品価格_再掲",
    "利益",
]

COLOR_SYSTEM_MAP = [
    # Order matters: more specific patterns first. Labels must match Buyma UI.
    (r"multi[\s\-]?color|multicolou?r|マルチ", "マルチカラー"),
    (r"navy|ネイビー|紺", "ネイビー（紺）系"),
    (r"black|nero|noir|黒|ブラック", "ブラック（黒）系"),
    (r"white|bianco|blanc|オフホワイト|ivory|アイボリー|白|ホワイト", "ホワイト（白）系"),
    (r"gray|grey|grigio|グレー|灰色|チャコール|charcoal", "グレー（灰色）系"),
    (r"brown|marrone|brun|茶|ブラウン|chocolate|チョコ", "ブラウン（茶色）系"),
    (r"beige|ベージュ|camel|キャメル|taupe|khaki|カーキ|sand|サンド", "ベージュ系"),
    (r"green|verde|vert|緑|グリーン|olive|オリーブ", "グリーン（緑）系"),
    (r"blue|blu|bleu|青|ブルー|azure|ライトブルー", "ブルー（青）系"),
    (r"purple|viola|紫|パープル|bordeaux|burgundy|ワイン", "パープル（紫）系"),
    (r"yellow|giallo|jaune|黄|イエロー|cream|crema|クリーム", "イエロー（黄色）系"),
    (r"pink|rosa|rose|ピンク", "ピンク系"),
    (r"red|rosso|rouge|赤|レッド", "レッド（赤）系"),
    (r"orange|オレンジ", "オレンジ系"),
    (r"gold|oro|ゴールド|金色", "ゴールド（金色）系"),
    (r"silver|argento|シルバー|銀色", "シルバー（銀色）系"),
    (r"clear|transparent|クリア|透明", "クリア（透明）系"),
]

_SYSTEM_TO_SHORT_NAME = {
    "ブラック（黒）系": "ブラック",
    "ホワイト（白）系": "ホワイト",
    "グレー（灰色）系": "グレー",
    "ブラウン（茶色）系": "ブラウン",
    "ベージュ系": "ベージュ",
    "グリーン（緑）系": "グリーン",
    "ブルー（青）系": "ブルー",
    "ネイビー（紺）系": "ネイビー",
    "パープル（紫）系": "パープル",
    "イエロー（黄色）系": "イエロー",
    "ピンク系": "ピンク",
    "レッド（赤）系": "レッド",
    "オレンジ系": "オレンジ",
    "ゴールド（金色）系": "ゴールド",
    "シルバー（銀色）系": "シルバー",
    "クリア（透明）系": "クリア",
    "マルチカラー": "マルチカラー",
}


def extract_raw_color(*texts: str | None) -> str:
    """Pick an explicit color token from Color:/カラー: lines or free text."""
    blob = "\n".join(t for t in texts if t)
    if not blob:
        return ""
    for pat in (
        r"(?:Color|Colour|Colore)\s*[:：]\s*([^\n|/]+)",
        r"カラー\s*[:：]\s*([^\n|/]+)",
        r"【カラー】\s*([^\n|/]+)",
    ):
        m = re.search(pat, blob, re.I)
        if m:
            raw = m.group(1).strip()
            if raw and raw not in {"その他", "指定なし", "na", "n/a", "-", "なし"}:
                return raw
    return ""


def infer_color_system(color: str | None, text: str | None = None) -> str:
    blob = f"{color or ''} {text or ''}"
    # Comment boilerplate includes ホワイトデー / バレンタイン — never treat as product color.
    blob = re.sub(r"ホワイトデー|バレンタインデー|バレンタイン", " ", blob)
    blob = blob.lower()
    if not blob.strip():
        return ""
    for pattern, label in COLOR_SYSTEM_MAP:
        if re.search(pattern, blob, re.I):
            return label
    return ""


_GARMENT_IN_COLOR_RE = re.compile(
    r"cardigan|jacket|skirt|boot|bag|shirt|pant|dress|short|earring|bracelet|"
    r"coat|sweater|blazer|jean|shoe|hat|scarf|hoodie|\btop\b|bermuda|sandal|"
    r"flat|pump|necklace",
    re.I,
)


def is_usable_color_name(name: str | None) -> bool:
    """True when ``name`` looks like a Buyma 色名, not a product title."""
    n = str(name or "").strip()
    if not n or n in {"その他", "指定なし", "na", "n/a", "-", "なし"}:
        return False
    if len(n) > 20:
        return False
    if re.search(r"[【】\[\]\"'「」]", n):
        return False
    if _GARMENT_IN_COLOR_RE.search(n):
        return False
    if re.search(r"関税|送料|レディース|ファッション|商品", n):
        return False
    if n in _SYSTEM_TO_SHORT_NAME:
        return False
    return True


def sanitize_color_name(system: str, *candidates: str | None) -> str:
    """Return a short Buyma 色名 that matches ``system`` (e.g. ブラウン)."""
    short = _SYSTEM_TO_SHORT_NAME.get(system, "")
    for raw in candidates:
        n = str(raw or "").strip()
        if not is_usable_color_name(n):
            continue
        inferred = infer_color_system(n)
        if inferred and system and inferred != system:
            continue
        if inferred == system and short:
            return short
        if n in _SYSTEM_TO_SHORT_NAME.values():
            return n
        return n[:20]
    return short or "ブラウン"


def resolve_buyma_color(*parts: str | None) -> tuple[str, str]:
    """Return (Buyma色の系統 label, 色名 input). Never maps unknown → マルチカラー."""
    cleaned = [str(p).strip() for p in parts if p and str(p).strip()]
    system = ""
    for p in cleaned:
        if p in _SYSTEM_TO_SHORT_NAME:
            system = p
            break
    if not system:
        raw = extract_raw_color(*cleaned)
        system = infer_color_system(raw, " ".join([raw] + cleaned))
    if not system:
        return "", ""
    name = sanitize_color_name(system, extract_raw_color(*cleaned), *cleaned)
    return system, name[:40]


def fetch_color_from_source_url(url: str) -> str:
    """Best-effort Color: extraction from EC product HTML (for already-scraped rows)."""
    u = (url or "").strip()
    if not u.startswith("http"):
        return ""
    try:
        import httpx

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            )
        }
        with httpx.Client(timeout=20.0, follow_redirects=True, headers=headers) as client:
            resp = client.get(u)
            if resp.status_code >= 400:
                return ""
            text = resp.text or ""
        m = re.search(
            r"(?:Color|Colour|Colore|カラー)\s*[:：]\s*([A-Za-zぁ-んァ-ン一-龥0-9 \-_/]+)",
            text,
            re.I,
        )
        if not m:
            # Julian often has plain text Color: Brown outside tags after strip
            m = re.search(r"Color:\s*([A-Za-z][A-Za-z \-/]{0,40})", text, re.I)
        if m:
            raw = re.sub(r"\s+", " ", m.group(1)).strip(" .-|/")
            if raw and raw.lower() not in {"na", "n/a", "other"}:
                return raw[:60]
    except Exception:  # noqa: BLE001
        return ""
    return ""


def resolve_listing_color(row: dict, *, fetch_color=None) -> tuple[str, str]:
    """Resolve color system + name for Engine3, refreshing from source URL when needed.

    Prefer CSV カラー系統 / 色. Do not scan 商品コメント (contains ホワイトデー etc.).
    """
    explicit_system = str(row.get("カラー系統") or "").strip()
    explicit_name = str(row.get("色") or "").strip()
    if explicit_system in {"その他", "指定なし"}:
        explicit_system = ""
    if explicit_name in {"その他", "指定なし"}:
        explicit_name = ""

    if explicit_system in _SYSTEM_TO_SHORT_NAME:
        name = sanitize_color_name(explicit_system, explicit_name)
        return explicit_system, name[:40]

    # Infer only from color fields + product/folder names — never comment/memo boilerplate.
    system, name = resolve_buyma_color(
        explicit_name,
        explicit_system,
        row.get("フォルダ名"),
        row.get("商品名"),
    )
    if system:
        return system, name or system

    url = str(row.get("仕入先URL") or "")
    fetched = ""
    if callable(fetch_color):
        try:
            fetched = str(fetch_color(url) or "").strip()
        except Exception:  # noqa: BLE001
            fetched = ""
    if not fetched:
        fetched = fetch_color_from_source_url(url)
    if fetched:
        system2, name2 = resolve_buyma_color(fetched, row.get("商品名"), row.get("フォルダ名"))
        if system2:
            return system2, name2 or fetched
        return "", fetched
    return "", ""


@dataclass
class BuymaListingInput:
    brand_name: str
    product_name: str
    source_url: str
    shop_name: str
    price: float
    currency: str = "EUR"
    price_text: str | None = None
    reference_price: float | None = None
    product_code: str | None = None
    model_line: str | None = None
    category: str | None = None
    color: str | None = None
    size_text: str | None = None
    material: str | None = None
    origin_country: str | None = None
    source_description: str | None = None
    ai_comment: str | None = None
    external_product_id: str | None = None
    inventory: int = 1
    locked_product_jpy: int | None = None


@dataclass
class PriceBreakdown:
    product_jpy: int
    overseas_jpy: int
    pack_jpy: int
    domestic_jpy: int
    purchase_subtotal: int
    taxable: int
    duty: int
    consumption_tax: int
    cost_total: int
    profit: int
    profit_rate: float
    listing_price: int
    buyma_fee: int
    source_amount: float
    source_currency: str
    fx_rate: float
    overseas_source_amount: float
    calc_text: str


def buyma_folder_name(inp: BuymaListingInput) -> str:
    """Example: GIVENCHY_Minigonna_Cargo_in_Denim_blu_72658"""
    brand = slugify(inp.brand_name or "brand", separator="_").upper()
    # Prefer URL slug segment when present (matches sample folder style).
    raw_product = _product_slug_from_url(inp.source_url) or slugify(inp.product_name or "item", separator="_")
    product = _title_case_slug(raw_product)
    color_raw = (inp.color or "na").split("-")[0].split("/")[0].strip()
    color = slugify(color_raw, separator="_") or "na"
    pid = _product_id(inp)
    parts = [p for p in [brand, product, color, pid] if p]
    return "_".join(parts)[:180]


def _title_case_slug(slug: str) -> str:
    """Match sample style: Minigonna_Cargo_in_Denim (small words stay lower)."""
    small = {"in", "and", "of", "the", "with", "for", "to", "a", "an", "on", "at"}
    parts = []
    for part in (slug or "").split("_"):
        if not part:
            continue
        low = part.lower()
        parts.append(low if low in small else low.capitalize())
    return "_".join(parts) or "item"


def _product_slug_from_url(url: str | None) -> str:
    if not url:
        return ""
    # .../product/{id}/{brand}/{cat}/{slug}
    m = re.search(r"/product/\d+/[^/]+/[^/]+/([^/?#]+)", url, re.I)
    if m:
        return slugify(m.group(1), separator="_")
    m = re.search(r"/([^/]+?)(?:\.html)?(?:\?|#|$)", url.rstrip("/"))
    if m and m.group(1).lower() not in {"en-jp", "en", "product", "www"}:
        return slugify(m.group(1), separator="_")
    return ""


def _product_id(inp: BuymaListingInput) -> str:
    if inp.external_product_id:
        m = re.search(r"(\d+)$", inp.external_product_id)
        if m:
            return m.group(1)
        return slugify(inp.external_product_id, separator="_")
    if inp.source_url:
        m = re.search(r"/product/(\d+)/", inp.source_url)
        if m:
            return m.group(1)
    return ""


def title_product_name(brand_name: str, product_name: str) -> str:
    return f"【{brand_name}】{product_name} 関税無＆送料無料"


def infer_category(product_name: str, source_url: str = "", fallback: str | None = None) -> str:
    if fallback and "ファッション" in fallback:
        # Ensure bags/shoes fallbacks still expand to 3-level keys when possible.
        fb = fallback.strip()
        if fb in {
            "レディースファッション バッグ",
            "レディースファッション シューズ",
            "レディースファッション ワンピース",
        } or " > " in fb or "バッグ・カバン" in fb or "靴・シューズ" in fb:
            # Continue keyword refinement below using product name.
            pass
        elif "その他アウター" in fb:
            pass
        elif re.search(r"ボトムス|トップス|アウター", fb) and len(re.split(r"[\s/>＞]+", fb)) >= 3:
            return fb
    blob = f"{product_name} {source_url}".lower()
    if re.search(r"earring|ピアス|イヤリング", blob):
        return "レディースファッション ファッション小物 ピアス"
    if re.search(r"bracelet|ブレスレット", blob):
        return "レディースファッション ファッション小物 ブレスレット"
    if re.search(r"necklace|ネックレス", blob):
        return "レディースファッション ファッション小物 ネックレス"
    if re.search(r"skirt|minigonna|スカート", blob):
        return "レディースファッション ボトムス スカート"
    if re.search(r"bermuda|\bshorts?\b|ショートパンツ", blob) and not re.search(
        r"\bt-?shirts?\b|\btshirts?\b", blob
    ):
        return "レディースファッション ボトムス ショートパンツ"
    if re.search(r"dress|ワンピース", blob):
        return "レディースファッション ワンピース"
    if re.search(r"coat|trench|コート", blob):
        return "レディースファッション アウター コート"
    if re.search(r"blazer|jacket|blouson|giubbotto|ジャケット|ブレザー", blob):
        return "レディースファッション アウター ジャケット"
    if re.search(r"bag|handbag|tote|clutch|backpack|バッグ|ポーチ", blob):
        if re.search(r"tote|トート", blob):
            return "レディースファッション バッグ トートバッグ"
        if re.search(r"clutch|クラッチ", blob):
            return "レディースファッション バッグ クラッチ"
        if re.search(r"handbag|ハンドバッグ|trapeze|kelly|birkin", blob):
            return "レディースファッション バッグ ハンドバッグ"
        return "レディースファッション バッグ"
    if re.search(
        r"shoe|sneaker|sandal|loafer|boot|pump|heel|靴|gazelle|slingback|ballet|flat",
        blob,
    ):
        if re.search(r"sneaker|gazelle|スニーカー", blob):
            return "レディースファッション シューズ スニーカー"
        if re.search(r"boot|ブーツ", blob):
            return "レディースファッション シューズ ブーツ"
        if re.search(r"ballet|flat|バレー|フラット", blob):
            return "レディースファッション シューズ フラットシューズ"
        if re.search(r"pump|heel|slingback|パンプス", blob):
            return "レディースファッション シューズ パンプス"
        if re.search(r"sandal|サンダル", blob):
            return "レディースファッション シューズ サンダル"
        return "レディースファッション シューズ"
    if re.search(r"boot.*jean|jean|pant|trouser|パンツ|デニム", blob):
        return "レディースファッション ボトムス パンツ"
    if re.search(r"hoodie|\bsweat\b|sweatshirt|スウェット", blob):
        return "レディースファッション トップス パーカー・スウェット"
    if re.search(r"cardigan|sweater|knit|カーディガン|ニット", blob):
        return "レディースファッション トップス ニット・セーター"
    if re.search(r"polo|shirt|blouse|top|tee|t-shirt|タンク|トップス", blob):
        return "レディースファッション トップス Tシャツ・カットソー"
    return fallback or "レディースファッション アウター その他アウター"


def refine_listing_category(row: dict) -> str:
    """Re-infer weak CSV categories (その他アウター) from name + URL at list time."""
    current = str(row.get("カテゴリ") or row.get("category") or "").strip()
    name = str(row.get("商品名") or "")
    url = str(row.get("仕入先URL") or row.get("フォルダ名") or "")
    if (not current) or ("その他アウター" in current):
        return infer_category(name, url, fallback=current or None)
    return current


def fx_rate_for(currency: str) -> float:
    settings = get_settings()
    cur = (currency or "EUR").upper()
    if cur == "JPY":
        return 1.0
    if cur == "USD":
        return float(settings.usd_to_jpy_rate)
    if cur == "GBP":
        return float(settings.gbp_to_jpy_rate)
    return float(settings.eur_to_jpy_rate)


def to_jpy_amount(amount: float, currency: str) -> int:
    """Convert source currency to JPY (truncate toward zero, matching sheet examples)."""
    from decimal import Decimal, ROUND_DOWN

    rate = fx_rate_for(currency)
    raw = Decimal(str(amount or 0)) * Decimal(str(rate))
    return int(raw.to_integral_value(rounding=ROUND_DOWN))


def source_amount_from_product_jpy(product_jpy: float, currency: str) -> float:
    """品代(海外仕入) is always stored as JPY. Convert back to source currency for pricing."""
    cur = (currency or "EUR").upper()
    jpy = float(product_jpy or 0)
    if jpy <= 0:
        return 0.0
    if cur == "JPY":
        return jpy
    rate = float(fx_rate_for(cur) or 1.0)
    if rate <= 0:
        return jpy
    return jpy / rate


def build_color_size_note(inp: BuymaListingInput) -> str:
    return BUYMA_COLOR_SIZE_NOTE_TEMPLATE.format(
        brand_name=inp.brand_name,
        product_name=inp.product_name,
        color=(inp.color or "指定なし"),
    )


def _feature_bullets(ai_comment: str | None, inp: BuymaListingInput) -> str:
    text = (ai_comment or "").strip()
    # If AI already returned bullets / long body, keep as feature lines.
    if text and ("・" in text or "\n" in text or len(text) > 120):
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        out = []
        for ln in lines:
            if ln.startswith("・") or ln.startswith("-") or ln.startswith("*"):
                out.append(ln if ln.startswith("・") else "・" + ln.lstrip("-* "))
            else:
                out.append(f"・{ln}")
        return "\n".join(out[:8])
    bullets = []
    if text:
        bullets.append(f"・{text}")
    material = (inp.material or "").strip()
    if material and material != "確認中":
        bullets.append(f"・素材は{material}を採用")
    origin = (inp.origin_country or "").strip()
    if origin:
        bullets.append(f"・生産国：{origin}")
    color = (inp.color or "").strip()
    if color and color != "指定なし":
        bullets.append(f"・カラー：{color}")
    if not bullets:
        bullets.append(f"・{inp.brand_name}のアイテムです。詳細は商品ページをご確認ください。")
    return "\n".join(bullets)


def build_product_comment(inp: BuymaListingInput) -> str:
    raw = (inp.ai_comment or "").strip()
    # Already assembled delivery comment (re-export / retry).
    if raw.startswith("☆在庫がない") or "【あんしんプラス】" in raw[:800]:
        return raw
    features = _feature_bullets(inp.ai_comment, inp)
    material = (inp.material or "確認中").strip() or "確認中"
    origin = (inp.origin_country or "イタリア").strip() or "イタリア"
    color = (inp.color or "指定なし").strip() or "指定なし"
    size = (inp.size_text or "指定なし").strip().split("\n")[0] or "指定なし"
    style_note = (
        f"{inp.brand_name}の「{inp.product_name}」は、上質な佇まいと実用性を兼ね備えた一品です。"
        "トップスやアウターとの合わせ方次第で、カジュアルからスマートまで幅広く演出できます。"
    )
    return BUYMA_PRODUCT_COMMENT_TEMPLATE.format(
        brand_name=inp.brand_name,
        product_name=inp.product_name,
        features=features,
        material=material,
        origin_country=origin,
        color=color,
        size=size,
        style_note=style_note,
    )


def calculate_buyma_pricing(
    inp: BuymaListingInput,
    *,
    profit_rate: float | None = None,
) -> PriceBreakdown:
    """Compute listing price from the confirmed sheet rules.

    ① 品代 = source × FX (EUR×185.68)
    ② 海外送料 = EUR50 × FX → ¥9,284
    ③ 仕入合計 = 品代 + 海外送料
    ④ 課税価格 = 品代 × 60%
       - 課税価格 ≤ ¥16,666 → 免税 (関税・消費税 0)
       - otherwise 関税10% / 消費税10%
    ⑤ 関税 = 課税 × 10%  (0 if exempt)
    ⑥ 消費税 = (課税 + 関税) × 10%  (0 if exempt)
    ⑦ 国内送料 = ¥1,200
    ⑧ 原価合計 = 仕入合計 + 関税 + 消費税 + 国内送料
    ⑨ 利益 = max(原価 × 3%, ¥3,000)
    ⑩ 販売価格 = (原価 + 利益) ÷ 0.923  → 100円単位
       BUYMA手数料 = 販売 × 7.7%
    """
    settings = get_settings()
    cur = (inp.currency or "EUR").upper()
    amount = float(inp.price or 0)
    rate = fx_rate_for(cur)
    product_jpy = (
        int(inp.locked_product_jpy)
        if inp.locked_product_jpy is not None
        else to_jpy_amount(amount, cur)
    )

    overseas_eur = float(settings.buyma_overseas_shipping_eur)
    overseas_jpy = to_jpy_amount(overseas_eur, "EUR")
    pack_jpy = int(settings.buyma_packaging_jpy)
    domestic_jpy = int(settings.buyma_domestic_shipping_jpy)
    if profit_rate is None:
        profit_rate = float(settings.buyma_profit_rate)
    else:
        profit_rate = float(profit_rate)
    # Accept either 0.03 or 3 (%) style.
    if profit_rate > 1.0:
        profit_rate = profit_rate / 100.0
    fee_keep = float(settings.buyma_fee_keep_rate)
    taxable_ratio = float(settings.buyma_taxable_ratio)
    duty_free_max = int(settings.buyma_duty_free_taxable_jpy)
    min_profit = int(settings.buyma_min_profit_jpy)

    purchase_subtotal = product_jpy + overseas_jpy
    # Taxable base is 品代 only (not overseas shipping). Truncate like the sheet.
    taxable = int(product_jpy * taxable_ratio)
    duty_exempt = taxable <= duty_free_max
    if duty_exempt:
        duty = 0
        consumption_tax = 0
        tax_note = f"免税(課税価格≤¥{duty_free_max:,})"
    else:
        duty = int(taxable * 0.10)
        consumption_tax = int((taxable + duty) * 0.10)
        tax_note = "関税10%"
    cost_total = purchase_subtotal + duty + consumption_tax + domestic_jpy + pack_jpy
    profit = max(int(cost_total * profit_rate), min_profit if cost_total > 0 else 0)
    # Reverse BUYMA fee: sell = (cost + profit) / 0.923, nearest ¥100.
    listing_raw = (cost_total + profit) / fee_keep
    listing_price = int(math.ceil(listing_raw / 100.0 - 1e-9) * 100)
    if listing_price < cost_total:
        listing_price = int(round(listing_raw))
    buyma_fee = int(round(listing_price * (1.0 - fee_keep)))

    eur_rate = fx_rate_for("EUR")
    if cur == "JPY":
        product_line = f"① 商品(免税JPY): ¥{product_jpy:,}  (JPY{amount:,.2f})"
    else:
        product_line = f"① 商品(免税JPY): ¥{product_jpy:,}  ({cur}{amount:,.2f}×{rate:.2f})"
    fee_pct = (1.0 - fee_keep) * 100.0
    calc_text = "\n".join(
        [
            product_line,
            f"② 海外送料: ¥{overseas_jpy:,}  (EUR{overseas_eur:,.2f}×{eur_rate:.2f})",
            f"③ 仕入合計: ¥{purchase_subtotal:,}",
            f"④ 課税価格(×{taxable_ratio:.1f}): ¥{taxable:,}  [{tax_note}]",
            f"⑤ 関税 (10%): ¥{duty:,}",
            f"⑥ 消費税 (10%): ¥{consumption_tax:,}",
            f"⑦ 国内送料: ¥{domestic_jpy:,}",
            *( [f"⑦b 梱包: ¥{pack_jpy:,}"] if pack_jpy else [] ),
            f"⑧ 原価合計: ¥{cost_total:,}",
            f"⑨ 利益 ({profit_rate:.0%}): ¥{profit:,}",
            f"⑩ BUYMA手数料控除 (÷{fee_keep:.3f}): ¥{listing_price:,}",
            f"   BUYMA手数料: ¥{buyma_fee:,}",
        ]
    )

    return PriceBreakdown(
        product_jpy=product_jpy,
        overseas_jpy=overseas_jpy,
        pack_jpy=pack_jpy,
        domestic_jpy=domestic_jpy,
        purchase_subtotal=purchase_subtotal,
        taxable=taxable,
        duty=duty,
        consumption_tax=consumption_tax,
        cost_total=cost_total,
        profit=profit,
        profit_rate=profit_rate,
        listing_price=listing_price,
        buyma_fee=buyma_fee,
        source_amount=amount,
        source_currency=cur,
        fx_rate=rate,
        overseas_source_amount=overseas_eur,
        calc_text=calc_text,
    )


def build_buyma_row(
    inp: BuymaListingInput,
    *,
    profit_rate: float | None = None,
) -> dict[str, object]:
    settings = get_settings()
    folder = buyma_folder_name(inp)
    pricing = calculate_buyma_pricing(inp, profit_rate=profit_rate)
    comment = build_product_comment(inp)
    note = build_color_size_note(inp)
    sizes = normalize_buyma_sizes(inp.size_text)
    size_text = format_video_style_size_text(sizes)
    color_system, color_name = resolve_buyma_color(
        inp.color,
        inp.source_description,
        folder,
    )
    if not color_system:
        color_system, color_name = resolve_buyma_color(inp.color, folder)
    color_name = sanitize_color_name(color_system, color_name, inp.color)
    sku_parts: list[str] = []
    for part in (inp.product_code or "", inp.model_line or ""):
        p = str(part).strip()
        if p and p not in sku_parts and "①" not in p and "課税価格" not in p:
            sku_parts.append(p)
    sku = " ".join(sku_parts).strip()
    source_url = str(inp.source_url or "").strip()
    memo = "\n".join(
        [
            f"URL: {source_url}",
            f"SKU: {sku}" if sku else "SKU:",
            "",
            pricing.calc_text,
        ]
    )
    ref_jpy = ""
    if inp.reference_price:
        ref_jpy = to_jpy_amount(float(inp.reference_price), inp.currency)

    # Rebuild note with resolved color name when possible.
    if color_name:
        note_inp = BuymaListingInput(**{**inp.__dict__, "color": color_name})
        note = build_color_size_note(note_inp)

    return {
        "フォルダ名": folder,
        "商品名": title_product_name(inp.brand_name, inp.product_name),
        "ブランド": inp.brand_name,
        "モデル・ライン": inp.model_line or "",
        "カテゴリ": infer_category(inp.product_name, source_url, inp.category),
        "商品コメント": comment,
        "色・サイズ補足情報": note,
        "購入期限": "",
        "仕入先URL": source_url,
        "買付地": settings.buyma_procurement_area,
        "ショップ名": inp.shop_name or "",
        "発送地": settings.buyma_ship_from,
        "仕入先保存": inp.shop_name,
        "カラー系統": color_system or "",
        "色": color_name or (inp.color or ""),
        "サイズ": size_text,
        "シーズン": "",
        "タグ": "",
        "テーマ": "",
        "価格": pricing.listing_price,
        "参考価格": ref_jpy,
        "配送方法名": settings.buyma_shipping_method,
        "在庫": str(max(0, int(inp.inventory or 0))),
        "型番/メモ": sku,
        "関税負担": settings.buyma_duty_burden,
        "出品メモ": memo,
        "ライバル価格": "",
        "ライバルURL": "",
        "SKU": sku,
        "価格計算式": pricing.calc_text,
        "品代(海外仕入)": pricing.product_jpy,
        "送料(仕入)": pricing.overseas_jpy,
        "梱包": pricing.pack_jpy,
        "送料(出品)": pricing.domestic_jpy,
        "出品価格": pricing.listing_price,
        "関税(10%)": pricing.duty,
        "消費税(10%)": pricing.consumption_tax,
        "仕入合計": pricing.cost_total,
        "BUYMA手数料": pricing.buyma_fee,
        "利益%": pricing.profit_rate,
        "出品価格_再掲": pricing.listing_price,
        "利益": pricing.profit,
        "通貨": (inp.currency or "EUR").upper(),
    }


def reprice_row_from_cost(
    row: dict[str, object],
    *,
    profit_rate: float | None = None,
) -> dict[str, str]:
    """Rebuild price columns from stored 品代/currency using current settings + optional profit_rate."""
    from core.csv_schema import ensure_row

    out = ensure_row(row)
    try:
        product_jpy = float(str(out.get("品代(海外仕入)") or "0").replace(",", "") or 0)
    except ValueError:
        product_jpy = 0.0
    currency = (out.get("通貨") or "EUR").strip() or "EUR"
    # 品代 is always JPY in our schema — reverse to source currency for pricing.
    source_amount = source_amount_from_product_jpy(product_jpy, currency)
    try:
        inventory = int(str(out.get("在庫") or "1").replace(",", "").strip() or "1")
    except ValueError:
        inventory = 1
    inventory = max(0, inventory)
    sku_raw = str(out.get("型番/メモ") or out.get("SKU") or "")
    if "①" in sku_raw or "課税価格" in sku_raw:
        sku_raw = ""
    inp = BuymaListingInput(
        brand_name=out.get("ブランド") or "",
        product_name=(out.get("商品名") or "").replace("【", "").split("】")[-1].replace(" 関税無＆送料無料", "").strip()
        or (out.get("商品名") or ""),
        source_url=out.get("仕入先URL") or "",
        shop_name=out.get("仕入先保存") or out.get("ショップ名") or "",
        price=float(source_amount or 0),
        currency=currency,
        color=out.get("色") or out.get("カラー系統") or "",
        size_text=out.get("サイズ") or "",
        category=out.get("カテゴリ") or "",
        product_code=sku_raw,
        inventory=inventory,
        locked_product_jpy=int(product_jpy) if product_jpy else None,
    )
    priced = build_buyma_row(inp, profit_rate=profit_rate)
    for key in (
        "価格",
        "出品価格",
        "出品価格_再掲",
        "品代(海外仕入)",
        "送料(仕入)",
        "梱包",
        "送料(出品)",
        "関税(10%)",
        "消費税(10%)",
        "仕入合計",
        "BUYMA手数料",
        "利益%",
        "利益",
        "価格計算式",
        "通貨",
    ):
        if key in priced:
            out[key] = "" if priced[key] is None else str(priced[key])
    # Refresh memo calc block when present.
    memo = str(out.get("出品メモ") or "")
    calc = str(priced.get("価格計算式") or "")
    if calc and "① 商品" in memo:
        # Replace trailing calc text after blank line following SKU.
        parts = memo.split("\n\n", 1)
        head = parts[0]
        out["出品メモ"] = f"{head}\n\n{calc}".strip()
    elif calc and not memo.strip():
        out["出品メモ"] = calc
    return out


def apply_listing_defaults(row: dict[str, object]) -> dict[str, str]:
    """Fill empty Buyma listing fields from settings so Engine3 can list from CSV alone."""
    from datetime import datetime, timedelta

    from core.csv_schema import ensure_row
    from core.prompts.defaults import BUYMA_COLOR_SIZE_NOTE_TEMPLATE, BUYMA_PRODUCT_COMMENT_TEMPLATE

    settings = get_settings()
    out = ensure_row(row)
    defaults: dict[str, str] = {
        "買付地": settings.buyma_procurement_area,
        "発送地": settings.buyma_ship_from,
        "配送方法名": settings.buyma_shipping_method,
        "関税負担": settings.buyma_duty_burden,
        "在庫": "1",
        "サイズ": "指定なし",
        "カテゴリ": "レディースファッション > アウター > その他アウター",
    }
    for key, value in defaults.items():
        if not (out.get(key) or "").strip():
            out[key] = value
    # Shop display name: prefer explicit ショップ名, else 仕入先保存.
    if not (out.get("ショップ名") or "").strip() and (out.get("仕入先保存") or "").strip():
        out["ショップ名"] = out["仕入先保存"]
    if not (out.get("買付先ショップ名") or "").strip():
        out["買付先ショップ名"] = (out.get("ショップ名") or out.get("仕入先保存") or "").strip()
    # Keep price columns aligned.
    if not (out.get("価格") or "").strip() and (out.get("出品価格") or "").strip():
        out["価格"] = out["出品価格"]
    if not (out.get("出品価格") or "").strip() and (out.get("価格") or "").strip():
        out["出品価格"] = out["価格"]
    # Purchase deadline required on current Buyma UI.
    if not (out.get("購入期限") or "").strip():
        out["購入期限"] = (datetime.now() + timedelta(days=14)).strftime("%Y/%m/%d")

    # Resolve real color (never keep placeholder その他 → マルチカラー / product titles).
    cur_system = (out.get("カラー系統") or "").strip()
    cur_color = (out.get("色") or "").strip()
    if cur_system in _SYSTEM_TO_SHORT_NAME and not is_usable_color_name(cur_color):
        out["色"] = sanitize_color_name(cur_system, cur_color)
    elif cur_system in {"", "その他"} or not cur_color:
        system, color_name = resolve_listing_color(out)
        if system:
            out["カラー系統"] = system
        if color_name:
            out["色"] = color_name

    # Resolve sizes when blank / 指定なし (Engine3 will also refresh from source URL).
    cur_size = (out.get("サイズ") or "").strip()
    if cur_size in {"", "指定なし", "なし"}:
        sizes = resolve_listing_sizes(out)
        if sizes:
            out["サイズ"] = format_video_style_size_text(sizes)

    brand = (out.get("ブランド") or "").strip() or "ブランド"
    name = (out.get("商品名") or "").strip() or "商品"
    color = (out.get("色") or out.get("カラー系統") or "指定なし").strip() or "指定なし"
    size = (out.get("サイズ") or "指定なし").strip() or "指定なし"
    # Refresh color line inside note when it still says その他.
    note = (out.get("色・サイズ補足情報") or "").strip()
    if note and "【カラー】その他" in note and color not in {"", "その他", "指定なし"}:
        out["色・サイズ補足情報"] = note.replace("【カラー】その他", f"【カラー】{color}", 1)
    if not (out.get("商品コメント") or "").strip():
        try:
            out["商品コメント"] = BUYMA_PRODUCT_COMMENT_TEMPLATE.format(
                brand_name=brand,
                product_name=name,
                features=f"・{brand}のアイテムです。詳細は商品ページをご確認ください。",
                material="確認中",
                origin_country="イタリア",
                color=color,
                size=size,
                style_note=f"{brand}の「{name}」は、上質な佇まいと実用性を兼ね備えた一品です。",
            )
        except Exception:  # noqa: BLE001
            out["商品コメント"] = f"{brand} / {name}\nご注文前に在庫確認のお問い合わせをお願いいたします。"
    if not (out.get("色・サイズ補足情報") or "").strip():
        try:
            out["色・サイズ補足情報"] = BUYMA_COLOR_SIZE_NOTE_TEMPLATE.format(
                brand_name=brand,
                product_name=name,
                color=color,
            )
        except Exception:  # noqa: BLE001
            out["色・サイズ補足情報"] = f"カラー：{color}\nサイズ：{size}"

    # Keep 出品メモ URL line identical to 仕入先URL (canonical source link).
    source_url = (out.get("仕入先URL") or "").strip()
    if source_url:
        memo = (out.get("出品メモ") or "").strip()
        if memo:
            lines = memo.splitlines()
            replaced = False
            for i, line in enumerate(lines):
                if line.strip().lower().startswith("url:"):
                    lines[i] = f"URL: {source_url}"
                    replaced = True
                    break
            if not replaced:
                lines = [f"URL: {source_url}", *lines]
            out["出品メモ"] = "\n".join(lines)
        else:
            out["出品メモ"] = f"URL: {source_url}"
    return out


def write_listing_csv_beside_images(folder: Path, row: dict[str, object]) -> Path:
    """Persist listing fields next to product images as CSV only (no TXT sidecars)."""
    from core.csv_schema import write_products_csv

    folder = Path(folder)
    folder.mkdir(parents=True, exist_ok=True)
    ready = apply_listing_defaults(row)
    path = folder / "buyma_listing.csv"
    write_products_csv(path, [ready])
    return path


def row_values_for_excel(row: dict[str, object]) -> list[object]:
    return [row.get(k, "") for k in _ROW_KEYS]


def save_buyma_production_files(
    target: Path,
    inp: BuymaListingInput,
    *,
    description_prompt: str = "",
    image_prompt: str | None = None,
    append_delivery_sheet: bool = True,
) -> dict[str, Path]:
    """Save Buyma listing CSV beside images (+ optional delivery xlsx). Text stays in CSV columns."""
    target.mkdir(parents=True, exist_ok=True)
    row = apply_listing_defaults(build_buyma_row(inp))
    csv_path = write_listing_csv_beside_images(target, row)
    paths: dict[str, Path] = {"csv": csv_path}
    if append_delivery_sheet:
        xlsx_path = append_delivery_excel_row(row)
        paths["delivery_xlsx"] = xlsx_path
    return paths


def delivery_xlsx_path(*, run_id: int | None = None, day: datetime | None = None) -> Path:
    settings = get_settings()
    root = settings.export_root / "delivery"
    root.mkdir(parents=True, exist_ok=True)
    day = day or datetime.now()
    if run_id:
        return root / f"ご納品シート_{day.strftime('%Y%m%d')}_run{run_id}.xlsx"
    return root / f"ご納品シート_{day.strftime('%Y%m%d')}.xlsx"


def append_delivery_excel_row(row: dict[str, object], *, path: Path | None = None) -> Path:
    """Append one product row to the daily delivery workbook (sample columns)."""
    xlsx = path or delivery_xlsx_path()
    values = row_values_for_excel(row)
    with _delivery_lock:
        if xlsx.exists():
            wb = load_workbook(xlsx)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "ご納品シート"
            ws.append(BUYMA_CSV_HEADERS)
        # Avoid duplicate folder rows if pipeline retries the same item.
        folder = str(row.get("フォルダ名") or "")
        if folder:
            for existing in ws.iter_rows(min_row=2, max_col=1, values_only=False):
                if existing[0].value == folder:
                    r = existing[0].row
                    for idx, val in enumerate(values, start=1):
                        ws.cell(r, idx, val)
                    wb.save(xlsx)
                    logger.info("Updated delivery row for %s -> %s", folder, xlsx)
                    return xlsx
        ws.append(values)
        wb.save(xlsx)
    logger.info("Appended delivery row for %s -> %s", folder or "?", xlsx)
    return xlsx


def extract_color_from_notes(notes: str | None) -> str | None:
    if not notes:
        return None
    m = re.search(r"Color:\s*([^\n]+)", notes, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r"カラー[：:]\s*([^\n]+)", notes)
    if m:
        return m.group(1).strip()
    return None


def extract_sizes_from_notes(notes: str | None) -> list[str]:
    if not notes:
        return []
    m = re.search(r"Sizes?:\s*([^\n]+)", notes, re.I)
    if not m:
        m = re.search(r"サイズ[：:]\s*([^\n]+)", notes)
    if not m:
        return []
    return normalize_buyma_sizes(m.group(1))


def normalize_buyma_sizes(raw: str | list[str] | None) -> list[str]:
    """Normalize EC/legacy size tokens into Buyma-friendly labels."""
    if raw is None:
        return []
    if isinstance(raw, str):
        # Legacy/video rows can be: 指定なし\nS,S\nM,M
        chunks = [ln.strip() for ln in str(raw).splitlines() if ln.strip()]
        parts: list[str] = []
        for ch in chunks:
            parts.extend(re.split(r"[,/、|]+", ch))
    else:
        parts = list(raw)
    out: list[str] = []
    onesize_tokens = {
        "U",
        "UNI",
        "UNIQUE",
        "ONE",
        "ONESIZE",
        "ONE SIZE",
        "OS",
        "FREE",
        "F",
        "フリー",
        "フリーサイズ",
        "指定なし",
        "なし",
    }
    for part in parts:
        s = str(part or "").strip()
        if not s:
            continue
        if s in {"指定なし", "なし"}:
            continue
        up = s.upper()
        if up in onesize_tokens or s in onesize_tokens:
            # Single free/onesize marker — Buyma often uses FREE SIZE / バリエーションなし
            label = "FREE SIZE"
        else:
            label = up if re.fullmatch(r"[A-Z0-9]{1,5}", up) else s
        if label not in out:
            out.append(label)
    return out


def format_video_style_size_text(sizes: list[str]) -> str:
    """Legacy/video style size cell format used by old System3 CSV."""
    norm = normalize_buyma_sizes(sizes)
    if not norm:
        return "指定なし"
    if sizes_prefer_no_variation(norm):
        return "指定なし\nフリー,"
    lines = ["指定なし"]
    for s in norm:
        lines.append(f"{s},{s}")
    return "\n".join(lines)


def listing_stock_qty(raw: str | int | None, sizes: list[str] | None = None) -> int:
    """Buyma 買付数量. Cap insane scrape totals (200+) to size-count."""
    digits = re.sub(r"[^\d]", "", str(raw or ""))
    try:
        n = int(digits or "1")
    except ValueError:
        n = 1
    size_n = len([s for s in (sizes or []) if str(s).strip()]) or 1
    if n > 20:
        n = size_n
    return min(max(n, 1), 99)


def sizes_prefer_no_variation(sizes: list[str]) -> bool:
    """True when product is onesize / free — Buyma バリエーションなし."""
    if not sizes:
        return True
    compact = {s.upper() for s in sizes}
    return compact <= {"FREE SIZE", "FREE", "F", "ONE SIZE", "U", "UNI", "OS"}


def resolve_listing_sizes(row: dict, *, fetch_sizes=None) -> list[str]:
    """Resolve available sizes for Engine3 (CSV first, then EC page fetch)."""
    existing = normalize_buyma_sizes(row.get("サイズ"))
    if existing and not sizes_prefer_no_variation(existing):
        return existing
    if existing and (row.get("サイズ") or "").strip() not in {"", "指定なし", "なし"}:
        # Explicit onesize already set
        return existing

    from_notes = extract_sizes_from_notes(str(row.get("出品メモ") or ""))
    if from_notes:
        return from_notes
    from_notes = extract_sizes_from_notes(str(row.get("色・サイズ補足情報") or ""))
    if from_notes:
        return from_notes

    url = str(row.get("仕入先URL") or "")
    fetched: list[str] = []
    if callable(fetch_sizes):
        try:
            raw = fetch_sizes(url)
            if isinstance(raw, list):
                fetched = normalize_buyma_sizes(raw)
            elif raw:
                fetched = normalize_buyma_sizes(str(raw))
        except Exception:  # noqa: BLE001
            fetched = []
    return fetched or existing


def extract_material_origin(text: str | None) -> tuple[str, str]:
    material = "確認中"
    origin = "イタリア"
    if not text:
        return material, origin
    m = re.search(r"(?:Composition|素材)[:：]\s*([^\n|]+)", text, re.I)
    if m:
        material = m.group(1).strip()
    elif re.search(r"100%\s*cotton|コットン", text, re.I):
        material = "100% コットン"
    elif re.search(r"100%\s*wool|ウール", text, re.I):
        material = "100% ウール"
    o = re.search(r"(?:Made In|生産国|made in)[:：\s]*([A-Za-zぁ-んァ-ン一-龥]+)", text, re.I)
    if o:
        raw = o.group(1).strip().split("|")[0]
        mapping = {
            "italy": "イタリア",
            "italia": "イタリア",
            "france": "フランス",
            "francia": "フランス",
            "spain": "スペイン",
            "uk": "イギリス",
            "usa": "アメリカ",
        }
        origin = mapping.get(raw.lower(), raw)
    return material, origin
