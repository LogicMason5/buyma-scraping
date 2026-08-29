"""Known-product registry for scrape dedupe (workbook + CSV history)."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from core.config import get_settings
from core.csv_schema import read_products_csv


def _norm_url(url: str) -> str:
    return (url or "").strip().split("?")[0].rstrip("/").lower()


def product_keys_from_row(row: dict) -> set[str]:
    keys: set[str] = set()
    ext = (row.get("外部商品ID") or "").strip()
    if ext:
        keys.add(f"id:{ext.lower()}")
    url = _norm_url(row.get("仕入先URL") or "")
    if url:
        keys.add(f"url:{url}")
    folder = (row.get("フォルダ名") or "").strip()
    if folder:
        keys.add(f"folder:{folder}")
    return keys


def load_known_product_keys(
    *,
    workbook_path: Path | None = None,
    extra_csv_paths: Iterable[Path] | None = None,
) -> set[str]:
    """Union of keys from the products workbook and optional CSV files."""
    settings = get_settings()
    known: set[str] = set()
    wb_path = Path(workbook_path or settings.products_workbook_path)
    if wb_path.exists():
        try:
            wb = load_workbook(wb_path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    rows = ws.iter_rows(values_only=True)
                    header = next(rows, None)
                    if not header:
                        continue
                    cols = {str(h).strip(): i for i, h in enumerate(header) if h}
                    for values in rows:
                        if not values:
                            continue
                        row = {
                            k: ("" if values[i] is None else str(values[i]))
                            for k, i in cols.items()
                            if i < len(values)
                        }
                        known |= product_keys_from_row(row)
            finally:
                wb.close()
        except Exception:  # noqa: BLE001
            pass

    for csv_path in extra_csv_paths or []:
        try:
            for row in read_products_csv(csv_path):
                known |= product_keys_from_row(row)
        except Exception:  # noqa: BLE001
            continue

    # Also scan recent scrape CSVs under workspace/scrape
    scrape_root = settings.workspace_dir / "scrape"
    if scrape_root.is_dir():
        for csv_path in sorted(scrape_root.glob("run_*/products.csv"))[-30:]:
            try:
                for row in read_products_csv(csv_path):
                    known |= product_keys_from_row(row)
            except Exception:  # noqa: BLE001
                continue

    ready = settings.workspace_dir / "generate" / "products_ready.csv"
    if ready.exists():
        try:
            for row in read_products_csv(ready):
                known |= product_keys_from_row(row)
        except Exception:  # noqa: BLE001
            pass

    return known


def scraped_product_keys(*, external_id: str = "", source_url: str = "", folder: str = "") -> set[str]:
    return product_keys_from_row(
        {
            "外部商品ID": external_id or "",
            "仕入先URL": source_url or "",
            "フォルダ名": folder or "",
        }
    )
