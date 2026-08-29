"""Local multi-sheet Excel workbook for batch product CSV management.

One file; sheet tabs are named ``YYYY-MM-DD (N)`` (today's date + row count).
Engine1 creates/appends today's sheet. Engine2 updates existing rows in place
(no new sheet).
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.csv_schema import ROW_KEYS, ensure_row

_DATE_SHEET_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})(?:\s*\((\d+)\))?$")


def today_date_label(day: datetime | None = None) -> str:
    """Sheet date prefix, e.g. ``2026-08-11`` (no slashes — invalid in Excel titles)."""
    day = day or datetime.now()
    return day.strftime("%Y-%m-%d")


def date_sheet_title(row_count: int, *, day: datetime | None = None) -> str:
    """``YYYY-MM-DD (N)`` — today's date only with count."""
    return f"{today_date_label(day)} ({max(0, int(row_count))})"


def _safe_sheet_title(name: str, *, used: set[str] | None = None) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]+", "-", (name or "").strip()) or "Sheet"
    cleaned = cleaned[:31]
    used = used or set()
    base = cleaned
    n = 2
    while cleaned in used:
        suffix = f"_{n}"
        cleaned = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(cleaned)
    return cleaned


def _open_or_create(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return load_workbook(path), False
    wb = Workbook()
    return wb, True


def _find_sheet_for_date(wb, day: datetime | None = None) -> Worksheet | None:
    label = today_date_label(day)
    for name in wb.sheetnames:
        if name == label or name.startswith(label + " "):
            return wb[name]
    return None


def _header_map(ws: Worksheet) -> dict[str, int]:
    """Column header -> 1-based index from row 1."""
    mapping: dict[str, int] = {}
    for col, cell in enumerate(ws[1], start=1):
        key = str(cell.value or "").strip()
        if key:
            mapping[key] = col
    return mapping


def _ensure_header_row(ws: Worksheet) -> dict[str, int]:
    mapping = _header_map(ws)
    if not (mapping.get("フォルダ名") and mapping.get("商品名")):
        # Write full schema header
        for col, key in enumerate(ROW_KEYS, start=1):
            ws.cell(1, col, key)
        return {k: i for i, k in enumerate(ROW_KEYS, start=1)}
    # Append any missing schema columns (e.g. 在庫監視 / Buyma公開) without rewriting.
    next_col = max(mapping.values(), default=0) + 1
    for key in ROW_KEYS:
        if key not in mapping:
            ws.cell(1, next_col, key)
            mapping[key] = next_col
            next_col += 1
    return mapping


def _data_row_count(ws: Worksheet) -> int:
    # Rows after header that have any value in col A or フォルダ名
    mapping = _header_map(ws)
    folder_col = mapping.get("フォルダ名", 1)
    count = 0
    for r in range(2, (ws.max_row or 1) + 1):
        val = ws.cell(r, folder_col).value
        if val is not None and str(val).strip() != "":
            count += 1
    return count


def _rename_sheet_with_count(wb, ws: Worksheet, *, day: datetime | None = None) -> str:
    n = _data_row_count(ws)
    desired = date_sheet_title(n, day=day)
    used = {s for s in wb.sheetnames if s != ws.title}
    title = _safe_sheet_title(desired, used=used)
    if ws.title != title:
        ws.title = title
    return title


def _write_row_values(ws: Worksheet, excel_row: int, row: dict[str, Any], mapping: dict[str, int]) -> None:
    ready = ensure_row(row)
    for key, col in mapping.items():
        if key in ready:
            ws.cell(excel_row, col, ready.get(key, ""))


def upsert_today_sheet(
    workbook_path: Path | str,
    rows: Sequence[dict[str, Any]],
    *,
    day: datetime | None = None,
) -> tuple[Path, str]:
    """Engine1: create today's sheet or append rows; sheet name ``YYYY-MM-DD (N)``."""
    path = Path(workbook_path)
    ready_rows = [ensure_row(r) for r in rows]
    wb, created = _open_or_create(path)
    day = day or datetime.now()

    if created:
        ws = wb.active
        assert ws is not None
        used: set[str] = set()
        title = _safe_sheet_title(date_sheet_title(len(ready_rows), day=day), used=used)
        ws.title = title
        for col, key in enumerate(ROW_KEYS, start=1):
            ws.cell(1, col, key)
        mapping = {k: i for i, k in enumerate(ROW_KEYS, start=1)}
        for i, row in enumerate(ready_rows):
            _write_row_values(ws, i + 2, row, mapping)
        title = _rename_sheet_with_count(wb, ws, day=day)
        wb.save(path)
        return path, title

    ws = _find_sheet_for_date(wb, day)
    if ws is None:
        title = _safe_sheet_title(date_sheet_title(len(ready_rows), day=day), used=set(wb.sheetnames))
        ws = wb.create_sheet(title=title)
        for col, key in enumerate(ROW_KEYS, start=1):
            ws.cell(1, col, key)

    mapping = _ensure_header_row(ws)
    folder_col = mapping.get("フォルダ名", 1)
    # Index existing folders on this sheet for update-vs-append
    existing: dict[str, int] = {}
    for r in range(2, (ws.max_row or 1) + 1):
        key = str(ws.cell(r, folder_col).value or "").strip()
        if key:
            existing[key] = r

    next_row = (ws.max_row or 1) + 1
    if next_row < 2:
        next_row = 2
    for row in ready_rows:
        folder = (row.get("フォルダ名") or "").strip()
        if folder and folder in existing:
            _write_row_values(ws, existing[folder], row, mapping)
        else:
            # Find first empty data row if max_row has blanks
            while next_row <= (ws.max_row or 1) and str(ws.cell(next_row, folder_col).value or "").strip():
                next_row += 1
            _write_row_values(ws, next_row, row, mapping)
            if folder:
                existing[folder] = next_row
            next_row += 1

    title = _rename_sheet_with_count(wb, ws, day=day)
    wb.save(path)
    return path, title


def update_workbook_rows(
    workbook_path: Path | str,
    rows: Sequence[dict[str, Any]],
    *,
    day: datetime | None = None,
) -> tuple[Path, str, int]:
    """Engine2: update existing rows by フォルダ名 (prefer today's sheet). No new sheet.

    Returns (path, sheet_title_touched, updated_count).
    """
    path = Path(workbook_path)
    ready_rows = [ensure_row(r) for r in rows]
    if not path.exists():
        # Nothing to update — create today's sheet once so data is not lost.
        p, title = upsert_today_sheet(path, ready_rows, day=day)
        return p, title, len(ready_rows)

    wb = load_workbook(path)
    day = day or datetime.now()
    today_ws = _find_sheet_for_date(wb, day)

    # Build folder -> (worksheet, excel_row) preferring today's sheet first
    loc: dict[str, tuple[Worksheet, int]] = {}
    sheets_order: list[Worksheet] = []
    if today_ws is not None:
        sheets_order.append(today_ws)
    sheets_order.extend(wb[name] for name in wb.sheetnames if today_ws is None or name != today_ws.title)

    for ws in sheets_order:
        mapping = _header_map(ws)
        folder_col = mapping.get("フォルダ名")
        if not folder_col:
            continue
        for r in range(2, (ws.max_row or 1) + 1):
            key = str(ws.cell(r, folder_col).value or "").strip()
            if key and key not in loc:
                loc[key] = (ws, r)

    updated = 0
    touched: Worksheet | None = today_ws
    for row in ready_rows:
        folder = (row.get("フォルダ名") or "").strip()
        if not folder:
            continue
        if folder in loc:
            ws, excel_row = loc[folder]
            mapping = _ensure_header_row(ws)
            _write_row_values(ws, excel_row, row, mapping)
            updated += 1
            touched = ws
        elif today_ws is not None:
            # New folder from Engine2 path: append on today's sheet
            mapping = _ensure_header_row(today_ws)
            folder_col = mapping.get("フォルダ名", 1)
            next_row = (today_ws.max_row or 1) + 1
            _write_row_values(today_ws, next_row, row, mapping)
            loc[folder] = (today_ws, next_row)
            updated += 1
            touched = today_ws

    title = ""
    if touched is not None:
        # If touched sheet is today's (or became the working sheet), refresh count in name
        if today_ws is not None and touched.title == today_ws.title:
            title = _rename_sheet_with_count(wb, touched, day=day)
        else:
            title = touched.title
            # Still refresh count if sheet follows date naming
            m = _DATE_SHEET_RE.match(touched.title)
            if m:
                d = datetime.strptime(m.group(1), "%Y-%m-%d")
                title = _rename_sheet_with_count(wb, touched, day=d)

    wb.save(path)
    return path, title or (today_ws.title if today_ws else ""), updated


# Back-compat aliases used by older call sites
def suggest_batch_sheet_name(*, run_id: str = "", row_count: int = 0, prefix: str = "") -> str:
    return date_sheet_title(row_count or 0)


def append_products_sheet(
    workbook_path: Path | str,
    rows: Sequence[dict[str, Any]],
    *,
    sheet_name: str = "",
    run_id: str = "",
) -> tuple[Path, str]:
    """Deprecated alias → upsert today's sheet."""
    return upsert_today_sheet(workbook_path, rows)


def list_workbook_sheets(workbook_path: Path | str) -> list[str]:
    path = Path(workbook_path)
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_all_workbook_rows(workbook_path: Path | str) -> list[dict[str, str]]:
    """Read every data row from all sheets (dedupe by フォルダ名/URL, newest sheet wins)."""
    path = Path(workbook_path)
    if not path.exists():
        return []
    by_folder: dict[str, dict[str, str]] = {}
    by_url: dict[str, dict[str, str]] = {}
    no_key: list[dict[str, str]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                continue
            cols = {str(h).strip(): i for i, h in enumerate(header) if h is not None and str(h).strip()}
            if "仕入先URL" not in cols and "フォルダ名" not in cols:
                continue
            for values in it:
                if not values:
                    continue
                row = {
                    k: ("" if (i >= len(values) or values[i] is None) else str(values[i]))
                    for k, i in cols.items()
                }
                folder = (row.get("フォルダ名") or "").strip()
                url = (row.get("仕入先URL") or "").strip().split("?")[0].rstrip("/").lower()
                ready = ensure_row(row)
                if folder:
                    by_folder[folder] = ready  # later sheets overwrite → prefer newest
                elif url:
                    by_url[url] = ready
                elif any(str(v).strip() for v in ready.values()):
                    no_key.append(ready)
    finally:
        wb.close()
    rows: list[dict[str, str]] = list(by_folder.values())
    seen_urls = {
        (r.get("仕入先URL") or "").strip().split("?")[0].rstrip("/").lower()
        for r in rows
        if (r.get("仕入先URL") or "").strip()
    }
    for url, row in by_url.items():
        if url and url not in seen_urls:
            rows.append(row)
            seen_urls.add(url)
    rows.extend(no_key)
    return rows


def patch_workbook_fields(
    workbook_path: Path | str,
    patches: Sequence[tuple[str, dict[str, Any]]],
) -> int:
    """Update only given fields for rows keyed by フォルダ名. Returns patched count.

    Does not blank unrelated cells — avoids wiping richer listing data during monitor runs.
    """
    path = Path(workbook_path)
    if not path.exists() or not patches:
        return 0
    # folder -> field map (last wins)
    by_folder: dict[str, dict[str, Any]] = {}
    for folder, fields in patches:
        key = (folder or "").strip()
        if not key or not fields:
            continue
        by_folder[key] = {**(by_folder.get(key) or {}), **fields}

    wb = load_workbook(path)
    updated = 0
    # Prefer today's sheet, then newer sheets first for locating the row.
    day = datetime.now()
    today_ws = _find_sheet_for_date(wb, day)
    sheets_order: list[Worksheet] = []
    if today_ws is not None:
        sheets_order.append(today_ws)
    sheets_order.extend(wb[name] for name in reversed(wb.sheetnames) if today_ws is None or name != today_ws.title)

    loc: dict[str, tuple[Worksheet, int]] = {}
    for ws in sheets_order:
        mapping = _header_map(ws)
        folder_col = mapping.get("フォルダ名")
        if not folder_col:
            continue
        for r in range(2, (ws.max_row or 1) + 1):
            key = str(ws.cell(r, folder_col).value or "").strip()
            if key and key not in loc:
                loc[key] = (ws, r)

    for folder, fields in by_folder.items():
        if folder not in loc:
            continue
        ws, excel_row = loc[folder]
        mapping = _ensure_header_row(ws)
        for key, value in fields.items():
            col = mapping.get(key)
            if not col:
                continue
            ws.cell(excel_row, col, "" if value is None else str(value))
        updated += 1

    if updated:
        wb.save(path)
    else:
        wb.close()
    return updated


def find_sheet_title_for_date(workbook_path: Path | str, day: datetime | None = None) -> str | None:
    """Return sheet title for ``YYYY-MM-DD`` (with optional `` (N)`` suffix), or None."""
    path = Path(workbook_path)
    if not path.exists():
        return None
    wb = load_workbook(path, read_only=True)
    try:
        label = today_date_label(day)
        for name in wb.sheetnames:
            if name == label or name.startswith(label + " "):
                return name
        return None
    finally:
        wb.close()


def resolve_sheet_name(workbook_path: Path | str, sheet_name: str) -> str | None:
    """Resolve exact sheet title, or same-date sheet when row-count suffix changed.

    Example: ``2026-08-11 (5)`` → ``2026-08-11 (10)`` if that is the live tab name.
    """
    wanted = (sheet_name or "").strip()
    if not wanted or wanted.startswith("("):
        return None
    path = Path(workbook_path)
    if not path.exists():
        return None
    wb = load_workbook(path, read_only=True)
    try:
        names = list(wb.sheetnames)
        if wanted in names:
            return wanted
        m = _DATE_SHEET_RE.match(wanted)
        if not m:
            return None
        label = m.group(1)
        for name in names:
            if name == label or name.startswith(label + " "):
                return name
        return None
    finally:
        wb.close()


def export_sheet_to_csv(
    workbook_path: Path | str,
    sheet_name: str,
    csv_path: Path | str,
) -> Path:
    """Export one workbook sheet to products CSV (ROW_KEYS aligned)."""
    from core.csv_schema import write_products_csv

    path = Path(workbook_path)
    out = Path(csv_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    resolved = resolve_sheet_name(path, sheet_name)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if not resolved or resolved not in wb.sheetnames:
            available = ", ".join(wb.sheetnames) or "(なし)"
            raise FileNotFoundError(
                f"シートが見つかりません: {sheet_name}\n利用可能: {available}"
            )
        ws = wb[resolved]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            write_products_csv(out, [])
            return out
        cols = {str(h).strip(): i for i, h in enumerate(header) if h}
        rows: list[dict[str, Any]] = []
        for values in rows_iter:
            if not values:
                continue
            row = {
                k: ("" if values[i] is None else str(values[i]))
                for k, i in cols.items()
                if i < len(values)
            }
            if not any(str(v).strip() for v in row.values()):
                continue
            rows.append(ensure_row(row))
        write_products_csv(out, rows)
        return out
    finally:
        wb.close()


def resolve_yesterday_sheet(workbook_path: Path | str, *, today: datetime | None = None) -> str | None:
    """Prefer yesterday's sheet; else newest prior date; else today's sheet."""
    from datetime import timedelta

    today = today or datetime.now()
    yesterday = today - timedelta(days=1)
    title = find_sheet_title_for_date(workbook_path, yesterday)
    if title:
        return title
    # Fallback: newest sheet whose date is strictly before today
    path = Path(workbook_path)
    if not path.exists():
        return None
    wb = load_workbook(path, read_only=True)
    try:
        best: tuple[datetime, str] | None = None
        for name in wb.sheetnames:
            m = _DATE_SHEET_RE.match(name)
            if not m:
                continue
            d = datetime.strptime(m.group(1), "%Y-%m-%d")
            if d.date() >= today.date():
                continue
            if best is None or d > best[0]:
                best = (d, name)
        if best:
            return best[1]
    finally:
        wb.close()
    # Same-day listing: only today's sheet exists yet.
    return find_sheet_title_for_date(workbook_path, today)

